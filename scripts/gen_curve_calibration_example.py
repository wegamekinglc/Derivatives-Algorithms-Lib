#!/usr/bin/env python3
"""
Generate dal-excel/examples/006.curve_calibration.xlsx

Mirrors dal-python/examples/004.curve_calibration.py for the Excel add-in:
an exactly-determined (square) OIS discount curve calibration using
deposits and OIS swaps, plus a second stage Libor 3M forward curve
calibration using FRAs and IRS.

Uses openpyxl to write the XLSX — run once, commit the .xlsx.
"""

import datetime
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

DST = os.path.join(
    os.path.dirname(__file__),
    "..",
    "dal-excel",
    "examples",
    "006.curve_calibration.xlsx",
)

wb = Workbook()

# ── style constants ──────────────────────────────────────────────────────────
TITLE_FONT = Font(name="Calibri", size=13, bold=True)
HEADER_FONT = Font(name="Calibri", size=11, bold=True)
BODY_FONT = Font(name="Calibri", size=11)
LABEL_FONT = Font(name="Calibri", size=10, color="555555")
MONO_FONT = Font(name="JetBrains Mono", size=10)  # fallback to Consolas
NUM_FMT = '#,##0.000000'
PCT_FMT = '0.000%'
BP_FMT = '0.00" bp"'
LIGHT_GRAY_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def _style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = LIGHT_GRAY_FILL


def _auto_width(ws, min_width=10, max_width=50):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


# =============================================================================
# Sheet 1: Single-curve OIS calibration
# =============================================================================
ws = wb.active
ws.title = "OIS Curve Calibration"
r = 1  # current row tracker

# ── Title ────────────────────────────────────────────────────────────────────
ws.cell(r, 1, "006.curve_calibration — Yield Curve Calibration").font = TITLE_FONT
r += 1
ws.cell(
    r,
    1,
    "Exactly-determined (square) OIS discount curve calibration. "
    "9 instruments on 9 knot pillars. "
    "The EXACT solver interpolates through every quote — residuals are at machine-epsilon level.",
).font = BODY_FONT
r += 2

# ── Setup ────────────────────────────────────────────────────────────────────
ws.cell(r, 1, "Setup").font = HEADER_FONT
r += 1
ws.cell(r, 1, "Init DAL").font = LABEL_FONT
ws.cell(r, 2, '=INIT.GLOBALDATA(0)').font = MONO_FONT
ws.cell(r, 3, "Run once before using any DAL functions")
r += 1
ws.cell(r, 1, "Evaluation Date").font = LABEL_FONT
ws.cell(r, 2, datetime.date(2024, 1, 15)).font = MONO_FONT
ws.cell(r, 2).number_format = "YYYY-MM-DD"
ws.cell(r, 3, "The valuation/trade date for all instruments")
r += 2

EVAL_DATE_CELL = "B6"  # row for evaluation date — used in instrument formulas
EVAL_DATE_ROW = 6

# ── Conventions ──────────────────────────────────────────────────────────────
ws.cell(r, 1, "Conventions").font = HEADER_FONT
r += 1
ws.cell(r, 1, "Overnight Index").font = LABEL_FONT
# RateIndexConventionNew(forecastTenor, basis, collateral, useProjectionCurve)
ws.cell(r, 2, '=RATEINDEXCONVENTION.NEW("1M","ACT_360","OIS",FALSE)').font = MONO_FONT
overnight_index_cell = f"B{r}"
r += 1
ws.cell(r, 1, "Libor 3M Index").font = LABEL_FONT
ws.cell(r, 2, '=RATEINDEXCONVENTION.NEW("3M","ACT_360","OIS",TRUE)').font = MONO_FONT
libor3m_index_cell = f"B{r}"
r += 1
ws.cell(r, 1, "Fixed Leg").font = LABEL_FONT
ws.cell(r, 2, '=RATELEGCONVENTION.NEW("6M","ACT_360")').font = MONO_FONT
fixed_leg_cell = f"B{r}"
r += 1
ws.cell(r, 1, "ON Float Leg").font = LABEL_FONT
ws.cell(r, 2, '=RATELEGCONVENTION.NEW("12M","ACT_360")').font = MONO_FONT
on_leg_cell = f"B{r}"
r += 1
ws.cell(r, 1, "Std Float Leg").font = LABEL_FONT
ws.cell(r, 2, '=RATELEGCONVENTION.NEW("6M","ACT_360")').font = MONO_FONT
float_leg_cell = f"B{r}"
r += 2

_style_header_row(ws, r - 6, 3)

# ── Stage 1: OIS discount curve ──────────────────────────────────────────────
ws.cell(r, 1, "Stage 1 — OIS Discount Curve").font = HEADER_FONT
r += 1
ws.cell(r, 1, "Instrument").font = HEADER_FONT
ws.cell(r, 2, "Handle").font = HEADER_FONT
ws.cell(r, 3, "Start").font = HEADER_FONT
ws.cell(r, 4, "Maturity").font = HEADER_FONT
ws.cell(r, 5, "Market Rate").font = HEADER_FONT
ws.cell(r, 6, "Knot Date").font = HEADER_FONT
_style_header_row(ws, r, 6)
r += 1

# Instrument rows
ois_instruments = [
    # (name, start_offset, maturity_offset, rate, knot_offset, formula_type)
    ("OIS Dep 1M", 0, 30, 0.01, 30, "deposit"),
    ("OIS Dep 3M", 0, 90, 0.01, 90, "deposit"),
    ("OIS Dep 6M", 0, 180, 0.01, 180, "deposit"),
    ("OIS Swp 12M", 0, 365, 0.01, 365, "ois"),
    ("OIS Swp 24M", 0, 730, 0.01, 730, "ois"),
    ("OIS Swp 36M", 0, 1095, 0.01, 1095, "ois"),
    ("OIS Swp 60M", 0, 1825, 0.01, 1825, "ois"),
    ("OIS Swp 84M", 0, 2555, 0.01, 2555, "ois"),
    ("OIS Swp 120M", 0, 3650, 0.01, 3650, "ois"),
]

ois_instrument_first_row = r
for i, (name, start_off, mat_off, rate, knot_off, typ) in enumerate(ois_instruments):
    ws.cell(r, 1, name).font = BODY_FONT
    # Raw expressions (no "=") for embedding inside function arguments.
    # A standalone cell formula starts with "="; an argument inside a formula does not.
    start_expr = EVAL_DATE_CELL if start_off == 0 else f"{EVAL_DATE_CELL}+{start_off}"
    mat_expr = f"{EVAL_DATE_CELL}+{mat_off}"
    knot_expr = f"{EVAL_DATE_CELL}+{knot_off}"

    if typ == "deposit":
        formula = (
            f'=DEPOSIT.NEW({EVAL_DATE_CELL},{start_expr},{mat_expr},'
            f'{rate},{overnight_index_cell})'
        )
    else:
        formula = (
            f'=OISSWAP.NEW({EVAL_DATE_CELL},{start_expr},{mat_expr},'
            f'{rate},{fixed_leg_cell},{overnight_index_cell},{on_leg_cell})'
        )

    ws.cell(r, 2, formula).font = MONO_FONT
    ws.cell(r, 3, f"={start_expr}").font = MONO_FONT
    ws.cell(r, 3).number_format = "YYYY-MM-DD"
    ws.cell(r, 4, f"={mat_expr}").font = MONO_FONT
    ws.cell(r, 4).number_format = "YYYY-MM-DD"
    ws.cell(r, 5, rate).font = MONO_FONT
    ws.cell(r, 5).number_format = PCT_FMT
    ws.cell(r, 6, f"={knot_expr}").font = MONO_FONT
    ws.cell(r, 6).number_format = "YYYY-MM-DD"
    r += 1

ois_instrument_last_row = r - 1
ois_instruments_range = f"B{ois_instrument_first_row}:B{ois_instrument_last_row}"
ois_knots_range = f"F{ois_instrument_first_row}:F{ois_instrument_last_row}"
r += 1

# Calibration settings table
ws.cell(r, 1, "Calibration Settings").font = HEADER_FONT
_style_header_row(ws, r, 2)
r += 1
settings_start = r
ois_settings = [
    ("curveName", "ois"),
    ("calibrateDiscountCurve", True),
    ("solveMode", "EXACT"),
    ("parameterization", "PIECEWISE_LINEAR_FWD"),
    ("targetCollateral", "OIS"),
]
for key, val in ois_settings:
    ws.cell(r, 1, key).font = MONO_FONT
    if isinstance(val, bool):
        ws.cell(r, 2, "TRUE" if val else "FALSE").font = MONO_FONT
    else:
        ws.cell(r, 2, val).font = MONO_FONT
    r += 1
settings_end = r - 1
ois_settings_range = f"A{settings_start}:B{settings_end}"
r += 1

# Calibration formula — returns a single result handle (curve + diagnostics bundled)
ws.cell(r, 1, "Calibrate").font = HEADER_FONT
calibration_formula = (
    f'=CALIBRATE.SINGLECURVE({EVAL_DATE_CELL},"USD",'
    f"{ois_instruments_range},{ois_knots_range},"
    f"{ois_settings_range})"
)
ws.cell(r, 2, calibration_formula).font = MONO_FONT
ws.cell(r, 3, "Returns a result handle; use CALIBRATIONRESULT.GET / .GET.CURVE to extract attributes")
ois_calib_row = r
r += 2

# ── Stage 2: Libor 3M forward curve ──────────────────────────────────────────
ws.cell(r, 1, "Stage 2 — Libor 3M Forward Curve").font = HEADER_FONT
r += 1
ws.cell(r, 1, "Instrument").font = HEADER_FONT
ws.cell(r, 2, "Handle").font = HEADER_FONT
ws.cell(r, 3, "Start").font = HEADER_FONT
ws.cell(r, 4, "Maturity").font = HEADER_FONT
ws.cell(r, 5, "Market Rate").font = HEADER_FONT
ws.cell(r, 6, "Knot Date").font = HEADER_FONT
_style_header_row(ws, r, 6)
r += 1

libor_instruments = [
    ("FRA 1x4", 30, 120, 0.03, 30, "fra"),
    ("FRA 3x6", 90, 180, 0.03, 90, "fra"),
    ("FRA 6x9", 180, 270, 0.03, 180, "fra"),
    ("IRS 12M", 0, 365, 0.03, 365, "swap"),
    ("IRS 24M", 0, 730, 0.03, 730, "swap"),
    ("IRS 36M", 0, 1095, 0.03, 1095, "swap"),
    ("IRS 60M", 0, 1825, 0.03, 1825, "swap"),
    ("IRS 84M", 0, 2555, 0.03, 2555, "swap"),
    ("IRS 120M", 0, 3650, 0.03, 3650, "swap"),
]

libor_instrument_first_row = r
for i, (name, start_off, mat_off, rate, knot_off, typ) in enumerate(libor_instruments):
    ws.cell(r, 1, name).font = BODY_FONT
    # Raw expressions (no "=") for embedding inside function arguments
    start_expr = EVAL_DATE_CELL if start_off == 0 else f"{EVAL_DATE_CELL}+{start_off}"
    mat_expr = f"{EVAL_DATE_CELL}+{mat_off}"
    knot_expr = f"{EVAL_DATE_CELL}+{knot_off}"

    if typ == "fra":
        formula = (
            f'=FRA.NEW({EVAL_DATE_CELL},{start_expr},{mat_expr},'
            f'{rate},{libor3m_index_cell})'
        )
    else:
        formula = (
            f'=SWAP.NEW({EVAL_DATE_CELL},{start_expr},{mat_expr},'
            f'{rate},{fixed_leg_cell},{libor3m_index_cell},{float_leg_cell})'
        )

    ws.cell(r, 2, formula).font = MONO_FONT
    ws.cell(r, 3, f"={start_expr}").font = MONO_FONT
    ws.cell(r, 3).number_format = "YYYY-MM-DD"
    ws.cell(r, 4, f"={mat_expr}").font = MONO_FONT
    ws.cell(r, 4).number_format = "YYYY-MM-DD"
    ws.cell(r, 5, rate).font = MONO_FONT
    ws.cell(r, 5).number_format = PCT_FMT
    ws.cell(r, 6, f"={knot_expr}").font = MONO_FONT
    ws.cell(r, 6).number_format = "YYYY-MM-DD"
    r += 1

libor_instrument_last_row = r - 1
libor_instruments_range = f"B{libor_instrument_first_row}:B{libor_instrument_last_row}"
libor_knots_range = f"F{libor_instrument_first_row}:F{libor_instrument_last_row}"
r += 1

# Calibration settings
ws.cell(r, 1, "Calibration Settings").font = HEADER_FONT
_style_header_row(ws, r, 2)
r += 1
settings_start2 = r
libor_settings = [
    ("curveName", "libor3m"),
    ("calibrateDiscountCurve", False),
    ("solveMode", "EXACT"),
    ("parameterization", "PIECEWISE_LINEAR_FWD"),
    ("targetCollateral", "OIS"),
    ("targetTenor", "3M"),
]
for key, val in libor_settings:
    ws.cell(r, 1, key).font = MONO_FONT
    if isinstance(val, bool):
        ws.cell(r, 2, "TRUE" if val else "FALSE").font = MONO_FONT
    else:
        ws.cell(r, 2, val).font = MONO_FONT
    r += 1
settings_end2 = r - 1
libor_settings_range = f"A{settings_start2}:B{settings_end2}"
r += 1

# Calibration — returns a single result handle (curve + diagnostics bundled).
# Stage 2 calibrates a FORWARD curve, so it needs the Stage-1 OIS discount curve
# as its discountCurve input (mirrors Python's discountCurves_ = {OIS: ois_curve}).
ws.cell(r, 1, "Calibrate").font = HEADER_FONT
calibration_formula2 = (
    f'=CALIBRATE.SINGLECURVE({EVAL_DATE_CELL},"USD",'
    f"{libor_instruments_range},{libor_knots_range},"
    f"{libor_settings_range},CALIBRATIONRESULT.GET.CURVE(B{ois_calib_row}))"
)
ws.cell(r, 2, calibration_formula2).font = MONO_FONT
ws.cell(r, 3, "Forward-curve calibration: passes the Stage-1 OIS curve as the discount curve")
libor_calib_row = r
r += 2

# ── Diagnostics: extract attributes from the calibration result handle ───────
ws.cell(r, 1, "Stage 1 Results — OIS Curve").font = HEADER_FONT
r += 1
ws.cell(r, 1, "Output").font = HEADER_FONT
ws.cell(r, 2, "Formula").font = HEADER_FONT
ws.cell(r, 3, "Description").font = HEADER_FONT
_style_header_row(ws, r, 3)
r += 1
ois_result = f"B{ois_calib_row}"
diag_refs = [
    ("Curve Handle", f"=CALIBRATIONRESULT.GET.CURVE({ois_result})", "Calibrated discount curve handle"),
    ("Max Abs Residual", f"=CALIBRATIONRESULT.GET({ois_result},\"maxAbsResidual\")", "Scalar — max absolute rate residual (bp)"),
    ("RMS Residual", f"=CALIBRATIONRESULT.GET({ois_result},\"rmsResidual\")", "Scalar — root-mean-square residual (bp)"),
    ("Residuals", f"=CALIBRATIONRESULT.GET({ois_result},\"residuals\")", "Column — (model - market) per instrument"),
    ("Market Rates", f"=CALIBRATIONRESULT.GET({ois_result},\"marketRates\")", "Column — quoted rate per instrument"),
    ("Model Rates", f"=CALIBRATIONRESULT.GET({ois_result},\"modelRates\")", "Column — model-implied rate per instrument"),
]
for label, formula, desc in diag_refs:
    ws.cell(r, 1, label).font = BODY_FONT
    ws.cell(r, 2, formula).font = MONO_FONT
    ws.cell(r, 3, desc).font = BODY_FONT
    r += 1

# Stage 2 (Libor) results
ws.cell(r, 1, "Stage 2 Results — Libor 3M Forward Curve").font = HEADER_FONT
r += 1
libor_result = f"B{libor_calib_row}"
diag_refs2 = [
    ("Curve Handle", f"=CALIBRATIONRESULT.GET.CURVE({libor_result})", "Calibrated forward curve handle"),
    ("Max Abs Residual", f"=CALIBRATIONRESULT.GET({libor_result},\"maxAbsResidual\")", "Scalar — max absolute rate residual (bp)"),
    ("RMS Residual", f"=CALIBRATIONRESULT.GET({libor_result},\"rmsResidual\")", "Scalar — root-mean-square residual (bp)"),
]
for label, formula, desc in diag_refs2:
    ws.cell(r, 1, label).font = BODY_FONT
    ws.cell(r, 2, formula).font = MONO_FONT
    ws.cell(r, 3, desc).font = BODY_FONT
    r += 1

_auto_width(ws)

# =============================================================================
# Sheet 2: Notes / Expected output
# =============================================================================
ws2 = wb.create_sheet("Notes")
r2 = 1

ws2.cell(r2, 1, "Expected Behaviour").font = TITLE_FONT
r2 += 2
notes = [
    "1. Open the file, enable macros if prompted (DAL add-in must be loaded).",
    "2. Excel will recalculate formulas — calibration results appear automatically.",
    "3. The calibration is exactly-determined (square): 9 instruments for 9 knot pillars.",
    "4. With the EXACT solver, residuals should be at machine-epsilon (~1e-15) level.",
    "5. Try modifying market rates or adding/removing instruments to experiment.",
    "",
    "The two stages work independently:",
    "  Stage 1 calibrates an OIS discount curve from deposits and OIS swaps.",
    "  Stage 2 calibrates a Libor 3M forward curve from FRAs and IRS.",
    "",
    "CALIBRATE.SINGLECURVE returns a single result handle bundling the curve and diagnostics.",
    "Extract attributes with two helpers:",
    "  CALIBRATIONRESULT.GET.CURVE(result)              -> calibrated discount curve handle",
    "  CALIBRATIONRESULT.GET(result, \"maxAbsResidual\")  -> scalar stat (also rmsResidual)",
    "  CALIBRATIONRESULT.GET(result, \"residuals\")      -> per-instrument column",
    "    (also marketRates, modelRates)",
    "",
    "Settings keys supported by CALIBRATE.SINGLECURVE:",
    "  curveName, calibrateDiscountCurve, solveMode, parameterization,",
    "  logDfScheme, smoothingWeight, tolerance, fitTolerance,",
    "  maxEvaluations, maxRestarts, initialGuess,",
    "  targetCollateral, targetTenor, liborBasis.",
]
for note in notes:
    ws2.cell(r2, 1, note).font = BODY_FONT
    r2 += 1

_auto_width(ws2, min_width=60)

# ── Save ─────────────────────────────────────────────────────────────────────
os.makedirs(os.path.dirname(DST), exist_ok=True)
wb.save(DST)
print(f"Wrote {DST}")
