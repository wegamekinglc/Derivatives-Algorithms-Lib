#!/usr/bin/env python3
"""
Generate dal-excel/examples/007.xccy_calibration.xlsx

Mirrors dal-python/examples/006.xccy_calibration.py for the Excel add-in:
cross-currency basis calibration of a USD/EUR market.

Uses openpyxl to write the XLSX — run once, commit the .xlsx.
"""

import datetime
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DST = os.path.join(
    os.path.dirname(__file__),
    "..",
    "dal-excel",
    "examples",
    "007.xccy_calibration.xlsx",
)

wb = Workbook()

# ── style constants ──────────────────────────────────────────────────────────
TITLE_FONT = Font(name="Calibri", size=13, bold=True)
HEADER_FONT = Font(name="Calibri", size=11, bold=True)
BODY_FONT = Font(name="Calibri", size=11)
LABEL_FONT = Font(name="Calibri", size=10, color="555555")
MONO_FONT = Font(name="JetBrains Mono", size=10)
PCT_FMT = '0.000%'
NUM_FMT = '0.000000'
GRAY_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")


def _style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = GRAY_FILL


def _auto_width(ws, min_width=10, max_width=60):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


# =============================================================================
# Sheet: XCCY Calibration
# =============================================================================
ws = wb.active
ws.title = "XCCY Calibration"
r = 1

ws.cell(r, 1, "007.xccy_calibration — Cross-Currency Basis Calibration").font = TITLE_FONT
r += 1
ws.cell(
    r,
    1,
    "Calibrate the USD/EUR cross-currency basis curve from a single basis swap. "
    "Exactly-determined: 1 instrument pins 1 free forward segment (2 knots). "
    "Residuals should be at machine-epsilon level.",
).font = BODY_FONT
r += 2

# ── Setup ────────────────────────────────────────────────────────────────────
ws.cell(r, 1, "Setup").font = HEADER_FONT
r += 1
ws.cell(r, 1, "Init DAL").font = LABEL_FONT
ws.cell(r, 2, "=INIT.GLOBALDATA(0)").font = MONO_FONT
ws.cell(r, 3, "Run once before using any DAL functions")
r += 1
ws.cell(r, 1, "Evaluation Date").font = LABEL_FONT
ws.cell(r, 2, datetime.date(2024, 1, 15)).font = MONO_FONT
ws.cell(r, 2).number_format = "YYYY-MM-DD"
ws.cell(r, 3, "The valuation/trade date")
EVAL = f"B{r}"  # B6
r += 2

# ── Flat OIS discount curves (USD 2%, EUR 1%) ────────────────────────────────
ws.cell(r, 1, "Flat OIS Discount Curves").font = HEADER_FONT
r += 1
ws.cell(r, 1, "Tenor").font = HEADER_FONT
ws.cell(r, 2, "Knot Date").font = HEADER_FONT
ws.cell(r, 3, "USD fwd").font = HEADER_FONT
ws.cell(r, 4, "EUR fwd").font = HEADER_FONT
_style_header_row(ws, r, 4)
r += 1

tenors = [365, 730, 1095, 1825, 3650, 7300, 10950]  # 1Y,2Y,3Y,5Y,10Y,20Y,30Y
usd_rate = 0.02
eur_rate = 0.01
knot_first = r
for t in tenors:
    ws.cell(r, 1, f"{t//365}Y").font = BODY_FONT
    ws.cell(r, 2, f"={EVAL}+{t}").font = MONO_FONT
    ws.cell(r, 2).number_format = "YYYY-MM-DD"
    ws.cell(r, 3, usd_rate).font = MONO_FONT
    ws.cell(r, 3).number_format = PCT_FMT
    ws.cell(r, 4, eur_rate).font = MONO_FONT
    ws.cell(r, 4).number_format = PCT_FMT
    r += 1
knot_last = r - 1
usd_knots = f"B{knot_first}:B{knot_last}"
usd_fwds = f"C{knot_first}:C{knot_last}"
eur_fwds = f"D{knot_first}:D{knot_last}"
r += 1

ws.cell(r, 1, "USD curve").font = LABEL_FONT
ws.cell(r, 2, f'=DISCOUNTPWLF.NEW("USD","USD",{usd_knots},{usd_fwds})').font = MONO_FONT
usd_curve = f"B{r}"
r += 1
ws.cell(r, 1, "EUR curve").font = LABEL_FONT
ws.cell(r, 2, f'=DISCOUNTPWLF.NEW("EUR","EUR",{usd_knots},{eur_fwds})').font = MONO_FONT
eur_curve = f"B{r}"
r += 1
ws.cell(r, 1, "USD block").font = LABEL_FONT
ws.cell(r, 2, f'=CURVEBLOCK.NEW.SIMPLE({usd_curve},"ACT_365F")').font = MONO_FONT
usd_block = f"B{r}"
r += 1
ws.cell(r, 1, "EUR block").font = LABEL_FONT
ws.cell(r, 2, f'=CURVEBLOCK.NEW.SIMPLE({eur_curve},"ACT_365F")').font = MONO_FONT
eur_block = f"B{r}"
r += 2

# ── Conventions ──────────────────────────────────────────────────────────────
ws.cell(r, 1, "Conventions").font = HEADER_FONT
r += 1
ws.cell(r, 1, "Domestic index (12M)").font = LABEL_FONT
ws.cell(r, 2, '=RATEINDEXCONVENTION.NEW("12M","ACT_365F","OIS",TRUE)').font = MONO_FONT
dom_index = f"B{r}"
r += 1
ws.cell(r, 1, "Domestic leg (12M)").font = LABEL_FONT
ws.cell(r, 2, '=RATELEGCONVENTION.NEW("12M","ACT_365F")').font = MONO_FONT
dom_leg = f"B{r}"
r += 1
ws.cell(r, 1, "Foreign index (12M)").font = LABEL_FONT
ws.cell(r, 2, '=RATEINDEXCONVENTION.NEW("12M","ACT_365F","OIS",TRUE)').font = MONO_FONT
for_index = f"B{r}"
r += 1
ws.cell(r, 1, "Foreign leg (12M)").font = LABEL_FONT
ws.cell(r, 2, '=RATELEGCONVENTION.NEW("12M","ACT_365F")').font = MONO_FONT
for_leg = f"B{r}"
r += 1
ws.cell(r, 1, "Currency pair").font = LABEL_FONT
ws.cell(r, 2, '=CURRENCYPAIR.NEW("USD","EUR")').font = MONO_FONT
pair = f"B{r}"
r += 2

# ── Cross-currency basis swap (single 2Y instrument at 0 spread) ─────────────
ws.cell(r, 1, "Instrument").font = HEADER_FONT
r += 1
ws.cell(r, 1, "Handle").font = HEADER_FONT
ws.cell(r, 2, "Maturity").font = HEADER_FONT
ws.cell(r, 3, "Market spread").font = HEADER_FONT
_style_header_row(ws, r, 3)
r += 1
# CROSSCURRENCYSWAP.NEW(tradeDate,start,maturity,marketRate,currencies,
#                       domesticLeg,domesticIndex,foreignLeg,foreignIndex,
#                       [domesticNotional],[foreignNotional])
mat_expr = f"{EVAL}+720"  # 2Y
ws.cell(r, 1, (
    f'=CROSSCURRENCYSWAP.NEW({EVAL},{EVAL},{mat_expr},0,{pair},'
    f"{dom_leg},{dom_index},{for_leg},{for_index},110,100)"
)).font = MONO_FONT
ws.cell(r, 2, f"={mat_expr}").font = MONO_FONT
ws.cell(r, 2).number_format = "YYYY-MM-DD"
ws.cell(r, 3, 0.0).font = MONO_FONT
ws.cell(r, 3).number_format = PCT_FMT
inst_row = r
instruments = f"A{inst_row}"  # single-cell range for one instrument
r += 2

# ── Basis knot dates (1Y, 2Y → 1 free forward segment) ───────────────────────
ws.cell(r, 1, "Basis Knot Dates").font = HEADER_FONT
r += 1
basis_knot_first = r
ws.cell(r, 1, f"={EVAL}+365").font = MONO_FONT  # 1Y
ws.cell(r, 1).number_format = "YYYY-MM-DD"
r += 1
ws.cell(r, 1, f"={EVAL}+730").font = MONO_FONT  # 2Y
ws.cell(r, 1).number_format = "YYYY-MM-DD"
r += 1
basis_knot_last = r - 1
basis_knots = f"A{basis_knot_first}:A{basis_knot_last}"
r += 1

# ── Settings ─────────────────────────────────────────────────────────────────
ws.cell(r, 1, "Calibration Settings").font = HEADER_FONT
_style_header_row(ws, r, 2)
r += 1
settings_start = r
settings = [
    ("fxSpot", 1.10),
    ("smoothingWeight", 1.0),
    ("tolerance", 1e-10),
    ("fitTolerance", 1e-6),
]
for key, val in settings:
    ws.cell(r, 1, key).font = MONO_FONT
    ws.cell(r, 2, val).font = MONO_FONT
    r += 1
settings_end = r - 1
settings_range = f"A{settings_start}:B{settings_end}"
r += 1

# ── Calibrate ────────────────────────────────────────────────────────────────
ws.cell(r, 1, "Calibrate").font = HEADER_FONT
ws.cell(r, 2, (
    f'=CALIBRATE.XCCYMARKET({EVAL},"USD","EUR",{usd_block},{eur_block},'
    f"{instruments},{basis_knots},{settings_range})"
)).font = MONO_FONT
ws.cell(r, 3, "Returns a result handle; use XCCYCALIBRATIONRESULT.GET / .GET.BASISCURVE")
result_cell = f"B{r}"
r += 2

# ── Results ──────────────────────────────────────────────────────────────────
ws.cell(r, 1, "Results").font = HEADER_FONT
r += 1
ws.cell(r, 1, "Output").font = HEADER_FONT
ws.cell(r, 2, "Formula").font = HEADER_FONT
ws.cell(r, 3, "Description").font = HEADER_FONT
_style_header_row(ws, r, 3)
r += 1
results = [
    ("Basis Curve", f"=XCCYCALIBRATIONRESULT.GET.BASISCURVE({result_cell})",
     "Calibrated basis discount curve handle"),
    ("Max Abs Residual", f'=XCCYCALIBRATIONRESULT.GET({result_cell},"maxAbsResidual")',
     "Scalar — max absolute residual (bp)"),
    ("RMS Residual", f'=XCCYCALIBRATIONRESULT.GET({result_cell},"rmsResidual")',
     "Scalar — root-mean-square residual (bp)"),
    ("Market Spread", f'=XCCYCALIBRATIONRESULT.GET({result_cell},"marketRates")',
     "Quoted basis spread per instrument"),
    ("Model Spread", f'=XCCYCALIBRATIONRESULT.GET({result_cell},"modelRates")',
     "Model-implied basis spread per instrument"),
    ("Residuals", f'=XCCYCALIBRATIONRESULT.GET({result_cell},"residuals")',
     "(model - market) per instrument"),
]
for label, formula, desc in results:
    ws.cell(r, 1, label).font = BODY_FONT
    ws.cell(r, 2, formula).font = MONO_FONT
    ws.cell(r, 3, desc).font = BODY_FONT
    r += 1

_auto_width(ws)

# =============================================================================
# Sheet: Notes
# =============================================================================
ws2 = wb.create_sheet("Notes")
n = 1
ws2.cell(n, 1, "Expected Behaviour").font = TITLE_FONT
n += 2
notes = [
    "1. Open the file, enable macros if prompted (DAL add-in must be loaded).",
    "2. Excel recalculates — calibration results appear automatically.",
    "3. Exactly-determined: 1 basis swap pins 1 free forward segment (2 knots).",
    "4. With the EXACT solver, residuals should be at machine-epsilon (~1e-15).",
    "",
    "Setup mirrors dal-python/examples/006.xccy_calibration.py:",
    "  USD OIS 2.0%, EUR OIS 1.0%, FX spot EURUSD = 1.10.",
    "  Single 2Y cross-currency basis swap at 0 spread (110 USD / 100 EUR notional).",
    "",
    "CALIBRATE.XCCYMARKET returns a single result handle bundling the basis curve",
    "and fit diagnostics. Extract with:",
    "  XCCYCALIBRATIONRESULT.GET.BASISCURVE(result)        -> basis curve handle",
    '  XCCYCALIBRATIONRESULT.GET(result, "maxAbsResidual") -> scalar stat',
    '  XCCYCALIBRATIONRESULT.GET(result, "residuals")      -> per-instrument column',
    "    (also marketRates, modelRates, rmsResidual)",
    "",
    "Settings keys: fxSpot, fxForwardCollateral, smoothingWeight, tolerance,",
    "fitTolerance, initialGuess, maxEvaluations, maxRestarts, solveMode.",
]
for note in notes:
    ws2.cell(n, 1, note).font = BODY_FONT
    n += 1
_auto_width(ws2, min_width=60)

# ── Save ─────────────────────────────────────────────────────────────────────
os.makedirs(os.path.dirname(DST), exist_ok=True)
wb.save(DST)
print(f"Wrote {DST}")
